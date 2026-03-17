"""Excel I/O — read contacts, validate phone numbers, write logs."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class Contact:
    phone_number: str
    name: str = ""
    custom_fields: dict[str, str] = field(default_factory=dict)
    row_index: int = 0  # original row for resume capability


class ExcelHandler:
    def __init__(self, file_path: str):
        self._path = Path(file_path)

    def read_contacts(self) -> list[Contact]:
        """Read contacts from the Excel file.

        First row is treated as headers.  ``phone_number`` column is required.
        ``name`` column is optional.  All other columns become custom_fields.
        Duplicates (by phone number) are removed, keeping the first occurrence.
        """
        if not self._path.exists():
            raise FileNotFoundError(f"Contacts file not found: {self._path}")

        wb = load_workbook(self._path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=False))
        if len(rows) < 2:
            wb.close()
            return []

        # Parse headers
        headers: list[str] = []
        for cell in rows[0]:
            val = str(cell.value).strip().lower() if cell.value else ""
            headers.append(val)

        if "phone_number" not in headers:
            wb.close()
            raise ValueError("Excel file must contain a 'phone_number' column")

        phone_idx = headers.index("phone_number")
        name_idx = headers.index("name") if "name" in headers else None

        contacts: list[Contact] = []
        seen_phones: set[str] = set()

        for row_num, row in enumerate(rows[1:], start=2):
            raw_phone = str(row[phone_idx].value).strip() if row[phone_idx].value else ""
            normalized = self.validate_phone(raw_phone)
            if normalized is None:
                continue
            if normalized in seen_phones:
                continue
            seen_phones.add(normalized)

            name = ""
            if name_idx is not None and row[name_idx].value:
                name = str(row[name_idx].value).strip()

            custom: dict[str, str] = {}
            for i, header in enumerate(headers):
                if header in ("phone_number", "name") or not header:
                    continue
                val = row[i].value if i < len(row) else None
                if val is not None:
                    custom[header] = str(val).strip()

            contacts.append(Contact(
                phone_number=normalized,
                name=name,
                custom_fields=custom,
                row_index=row_num,
            ))

        wb.close()
        return contacts

    @staticmethod
    def validate_phone(phone: str) -> str | None:
        """Normalize and validate a phone number.

        - Strips spaces, dashes, parentheses
        - Keeps leading ``+``
        - Returns digits-only (with optional leading +) or ``None`` if invalid
        """
        if not phone:
            return None

        # Remove common formatting characters
        cleaned = re.sub(r"[\s\-\(\)\.]", "", phone)

        # Ensure it starts with + or digits
        if cleaned.startswith("+"):
            digits = cleaned[1:]
            if not digits.isdigit():
                return None
            if len(digits) < 7 or len(digits) > 15:
                return None
            return cleaned
        else:
            digits = re.sub(r"[^\d]", "", cleaned)
            if len(digits) < 7 or len(digits) > 15:
                return None
            return digits
